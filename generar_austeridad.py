"""
GENERADOR FORMATO AUSTERIDAD — Art. 10 LFAR — SADER
Fuentes: Cuenta Pública cierre 2025 (CSV) + SICOP corte actual 2026 (CSV)

USO EN GOOGLE COLAB:
    !python generar_austeridad.py
"""
import importlib, subprocess, sys as _sys
def _ensure(pkg):
    try: importlib.import_module(pkg)
    except ImportError:
        print(f"   Instalando {pkg}...")
        subprocess.check_call([_sys.executable, "-m", "pip", "install", pkg, "-q"])
_ensure("openpyxl"); _ensure("pandas")

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re, os
from datetime import datetime

# ══════════════════════════════════════════════════════════════════════
#  MAPA EXACTO DEL FORMATO DE REFERENCIA
#  (fila, tipo, concepto_A, partida)
# ══════════════════════════════════════════════════════════════════════
# tipo: 'grupo' = fila con nombre de concepto + fórmula SUM
#       'partida' = fila de datos con número de partida
#       'grupo1' = grupo con una sola partida (fórmula =+D_partida)
#       'fotocopiado' = grupo fotocopiado (suma D42+D43)
MAPA = [
    # fila, tipo,         label_A,                  partida, formula_D, formula_E
    (13, 'grupo',  'Alimentación',           None,  '=SUM(D14:D17)',     '=SUM(E14:E17)'),
    (14, 'part',   None,                     22102, None, None),
    (15, 'part',   None,                     22104, None, None),
    (16, 'part',   None,                     22106, None, None),
    (17, 'part',   None,                     38501, None, None),
    (18, 'grupo',  'Arrendamientos',         None,  '=SUM(D19:D27)',     '=SUM(E19:E27)'),
    (19, 'part',   None,                     32201, None, None),
    (20, 'part',   None,                     32301, None, None),  # Arrendamientos
    (21, 'part',   None,                     32302, None, None),
    (22, 'part',   None,                     32303, None, None),
    (23, 'part',   None,                     32503, None, None),
    (24, 'part',   None,                     32505, None, None),
    (25, 'part',   None,                     32601, None, None),
    (26, 'part',   None,                     32701, None, None),
    (27, 'part',   None,                     32903, None, None),
    (28, 'grupo',  'Bienes informáticos',    None,  '=SUM(D29:D31)',     '=SUM(E29:E31)'),
    (29, 'part',   None,                     21401, None, None),
    (30, 'part',   None,                     35301, None, None),
    (31, 'part',   None,                     51501, None, None),
    (32, 'grupo',  'Combustibles',           None,  '=SUM(D33:D36)',     '=SUM(E33:E36)'),
    (33, 'part',   None,                     26102, None, None),
    (34, 'part',   None,                     26103, None, None),
    (35, 'part',   None,                     26104, None, None),
    (36, 'part',   None,                     26105, None, None),
    (37, 'grupo1', 'Congresos',              None,  '=+D38',             '=SUM(E38)'),
    (38, 'part',   None,                     38301, None, None),
    (39, 'grupo1', 'Energía eléctrica',      None,  '=+D40',             '=+E40'),
    (40, 'part',   None,                     31101, None, None),
    (41, 'grupof', 'Fotocopiado **',         None,  '=+D42+D43',         '=+E42+E43'),
    (42, 'part',   None,                     32301, None, None),  # Fotocopiado (mismo número, split manual)
    (43, 'part',   None,                     33602, None, None),
    (44, 'grupo',  'Internet',               None,  '=SUM(D45:D46)',     '=SUM(E45:E46)'),
    (45, 'part',   None,                     31603, None, None),
    (46, 'part',   None,                     31701, None, None),
    (47, 'grupo1', 'Mobiliario',             None,  '=+D48',             '=+E48'),
    (48, 'part',   None,                     35201, None, None),
    (49, 'grupo1', 'Papelería',              None,  '=+D50',             '=+E50'),
    (50, 'part',   None,                     21101, None, None),
    (51, 'grupo',  'Pasajes',                None,  '=SUM(D52:D57)',     '=SUM(E52:E57)'),
    (52, 'part',   None,                     37101, None, None),
    (53, 'part',   None,                     37104, None, None),
    (54, 'part',   None,                     37106, None, None),
    (55, 'part',   None,                     37201, None, None),
    (56, 'part',   None,                     37204, None, None),
    (57, 'part',   None,                     37206, None, None),
    (58, 'grupo',  'Remodelación de oficinas',None, '=SUM(D59:D60)',     '=SUM(E59:E60)'),
    (59, 'part',   None,                     35101, None, None),
    (60, 'part',   None,                     62202, None, None),
    (61, 'grupo1', 'Telefonía',              None,  '=+D62',             '=+E62'),
    (62, 'part',   None,                     31401, None, None),
    (63, 'grupo',  'Viáticos',               None,  '=SUM(D64:D66)',     '=SUM(E64:E66)'),
    (64, 'part',   None,                     37501, None, None),
    (65, 'part',   None,                     37504, None, None),
    (66, 'part',   None,                     37602, None, None),
]

DENOMINACIONES = {
    22102: 'Productos alimenticios para personas derivado de la prestación de servicios públicos de carácter social y operativos',
    22104: 'Productos alimenticios para el personal en las instalaciones de las dependencias y entidades',
    22106: 'Productos alimenticios para el personal derivado de actividades extraordinarias o que presten servicios en turnos para atención continua',
    38501: 'Gastos para alimentación de servidores públicos de mando',
    32201: 'Arrendamiento de edificios y locales',
    32301: 'Arrendamiento de equipo y bienes informáticos',
    32302: 'Arrendamiento de mobiliario',
    32303: 'Arrendamiento de equipo de telecomunicaciones',
    32503: 'Arrendamiento de vehículos terrestres, aéreos, marítimos, lacustres y fluviales para servidores públicos de mando',
    32505: 'Arrendamiento de vehículos terrestres, aéreos, marítimos, lacustres y fluviales para labores en campo y de supervisión',
    32601: 'Arrendamiento de maquinaria y equipo',
    32701: 'Patentes, derechos de autor, regalías y otros',
    32903: 'Otros Arrendamientos',
    21401: 'Materiales y útiles consumibles para el procesamiento en equipos y bienes informáticos',
    35301: 'Mantenimiento y conservación de bienes informáticos',
    51501: 'Bienes informáticos',
    26102: 'Combustibles, lubricantes y aditivos para vehículos terrestres, aéreos, marítimos, lacustres y fluviales destinados a servicios públicos y la operación de programas públicos',
    26103: 'Combustibles, lubricantes y aditivos para vehículos terrestres, aéreos, marítimos, lacustres y fluviales destinados a servicios administrativos',
    26104: 'Combustibles, lubricantes y aditivos para vehículos terrestres, aéreos, marítimos, lacustres y fluviales asignados a servidores públicos',
    26105: 'Combustibles, lubricantes y aditivos para maquinaria, equipo de producción y servicios especializados',
    38301: 'Congresos y convenciones',
    31101: 'Servicio de energía eléctrica',
    33602: 'Otros servicios comerciales',
    31603: 'Servicios de internet',
    31701: 'Servicios de conducción de señales analógicas y digitales',
    35201: 'Mantenimiento y conservación de mobiliario y equipo de administración',
    21101: 'Materiales y útiles de oficina',
    37101: 'Pasajes aéreos nacionales para labores en campo y de supervisión',
    37104: 'Pasajes aéreos nacionales para servidores públicos de mando en el desempeño de comisiones y funciones oficiales',
    37106: 'Pasajes aéreos internacionales para servidores públicos en el desempeño de comisiones y funciones oficiales',
    37201: 'Pasajes terrestres nacionales para labores en campo y de supervisión',
    37204: 'Pasajes terrestres nacionales para servidores públicos de mando en el desempeño de comisiones y funciones oficiales',
    37206: 'Pasajes terrestres internacionales para servidores públicos en el desempeño de comisiones y funciones oficiales',
    35101: 'Mantenimiento y conservación de inmuebles para la prestación de servicios administrativos',
    62202: 'Mantenimiento y rehabilitación de edificaciones no habitacionales',
    31401: 'Servicio telefónico convencional',
    37501: 'Viáticos nacionales para labores en campo y de supervisión',
    37504: 'Viáticos nacionales para servidores públicos en el desempeño de funciones oficiales',
    37602: 'Viáticos en el extranjero para servidores públicos en el desempeño de comisiones y funciones oficiales',
}

# Grupos de filas de encabezado (para fórmula Total general)
FILAS_GRUPO = [13, 18, 28, 32, 37, 39, 41, 44, 47, 49, 51, 58, 61, 63]

# ══════════════════════════════════════════════════════════════════════
#  LECTURA DE DATOS
# ══════════════════════════════════════════════════════════════════════

def _leer_csv(path):
    try:    return pd.read_csv(path, encoding='utf-8', low_memory=False)
    except: return pd.read_csv(path, encoding='latin-1', low_memory=False)

def _prep(df):
    df = df.copy()
    df['P5'] = (df['CAPITULO'].astype(int)*10000 + df['CONCEPTO'].astype(int)*1000 +
                df['PARTIDA_GENERICA'].astype(int)*100 + df['PARTIDA_ESPECIFICA'].astype(int))
    df['SC'] = df['ID_UNIDAD'].astype(str).str.match(r'^\d{3}$')
    return df

def leer_cp2025(path):
    """
    CP 2025 — Sector Central = URs numéricas 3 dígitos + UR G00
    Columna: solo EJERCIDO (cierre definitivo)
    Split 32301: UR 513 = Arrendamientos | resto = Fotocopiado
    33602 Fotocopiado: se captura manualmente (dato de DGRMIS), no se puede
    derivar del CSV — el script deja 0 y el usuario lo ajusta.
    """
    print("   Leyendo Cuenta Pública 2025...")
    df = _prep(_leer_csv(path))
    # Sector Central = numéricas 3 dígitos + G00
    mask = df['SC'] | (df['ID_UNIDAD'] == 'G00')
    sc   = df[mask]
    base = dict(sc.groupby('P5')['EJERCIDO'].sum() / 1_000_000)

    # Split 32301: UR 513 = Arrendamientos, resto = Fotocopiado
    p32301 = sc[sc['P5'] == 32301]
    base[('32301','arrend')] = p32301[p32301['ID_UNIDAD'] == '513']['EJERCIDO'].sum() / 1_000_000
    base[('32301','foto')]   = p32301[p32301['ID_UNIDAD'] != '513']['EJERCIDO'].sum() / 1_000_000

    # 33602 Fotocopiado: no se puede derivar automáticamente del CSV
    # (el total SC incluye muchas URs que no son fotocopiado)
    # Se deja en 0 — el usuario ajusta si conoce el dato
    base[('33602','foto')] = 0.0

    print(f"   CP 2025: {len(sc['P5'].unique())} partidas (SC + G00)")
    return base

def leer_sicop2026(path):
    """
    SICOP 2026 — Sector Central = URs numéricas 3 dígitos + UR G00
    Columna: solo EJERCIDO (EJ_TRAMITE introduce registros aún no confirmados)
    Split 32301: UR 513 = Arrendamientos | resto = Fotocopiado
    33602 Fotocopiado: igual que CP2025, se deja en 0.
    """
    print("   Leyendo SICOP 2026...")
    df = _prep(_leer_csv(path))
    mask = df['SC'] | (df['ID_UNIDAD'] == 'G00')
    sc   = df[mask]
    base = dict(sc.groupby('P5')['EJERCIDO'].sum() / 1_000_000)

    p32301 = sc[sc['P5'] == 32301]
    base[('32301','arrend')] = p32301[p32301['ID_UNIDAD'] == '513']['EJERCIDO'].sum() / 1_000_000
    base[('32301','foto')]   = p32301[p32301['ID_UNIDAD'] != '513']['EJERCIDO'].sum() / 1_000_000
    base[('33602','foto')]   = 0.0

    print(f"   SICOP 2026: {len(sc['P5'].unique())} partidas (SC + G00)")
    return base

def detectar_corte(path):
    MESES = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
              'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12,
              'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
              'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12}
    DIAS = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}
    MESES_ES = {1:'enero',2:'febrero',3:'marzo',4:'abril',5:'mayo',6:'junio',
                7:'julio',8:'agosto',9:'septiembre',10:'octubre',11:'noviembre',12:'diciembre'}
    ABR3 = {1:'ENE',2:'FEB',3:'MAR',4:'ABR',5:'MAY',6:'JUN',
            7:'JUL',8:'AGO',9:'SEP',10:'OCT',11:'NOV',12:'DIC'}
    nom = os.path.basename(path).upper()
    m = re.search(r'(\d{1,2})[-_\s]([A-Z]+)[-_\s](\d{4})', nom)
    if m:
        mn = MESES.get(m.group(2))
        if mn:
            d = int(m.group(1))
            a = int(m.group(3))
            return {'dia':d,'mes':mn,'anio':a,'mes_es':MESES_ES[mn],'abr3':ABR3[mn]}
    hoy = datetime.today()
    return {'dia':DIAS[hoy.month],'mes':hoy.month,'anio':hoy.year,
            'mes_es':MESES_ES[hoy.month],'abr3':ABR3[hoy.month]}

# ══════════════════════════════════════════════════════════════════════
#  ESTILOS — exactos del formato de referencia
# ══════════════════════════════════════════════════════════════════════

ENC_FILL = 'A50021'   # color exacto de encabezados (FFA50021 sin alpha)
FMT_NUM  = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
FMT_PCT  = '0.00%'
FONT     = 'Calibri'
SZ       = 11

def _f(bold=False, sz=SZ): return Font(name=FONT, size=sz, bold=bold)
def _fill(hex6): return PatternFill('solid', fgColor=hex6)
def _a(h='general', v='top', wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_enc(c, val, ha='center', va='center', nf='General', wrap=False):
    """Estilo de encabezado (rojo oscuro, Calibri 11 bold, texto blanco via theme:0)"""
    c.value = val
    c.fill  = _fill(ENC_FILL)
    c.font  = Font(name=FONT, size=SZ, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    c.number_format = nf

def _set_grupo(c, val):
    """Fila de encabezado de concepto: Calibri 11 bold, sin fill, ha=None, va=top"""
    c.value = val
    c.font  = _f(bold=True)
    c.alignment = _a('left', 'top')

def _set_grupo_num(c, formula, nf=FMT_NUM):
    c.value = formula
    c.font  = _f(bold=True)
    c.alignment = _a('center', 'top')
    c.number_format = nf

def _set_part_num(c, val, nf=FMT_NUM):
    c.value = val
    c.font  = _f()
    c.alignment = _a('center', 'top')
    c.number_format = nf

# ══════════════════════════════════════════════════════════════════════
#  GENERACIÓN
# ══════════════════════════════════════════════════════════════════════

def generar_austeridad(ruta_cp, ruta_sicop, ruta_salida=None):
    ej25 = leer_cp2025(ruta_cp)
    ej26 = leer_sicop2026(ruta_sicop)
    c = detectar_corte(ruta_sicop)
    d, m_es, abr3, anio = c['dia'], c['mes_es'], c['abr3'], c['anio']
    print(f"   Corte: {d} de {m_es} de {anio}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"AUSTERIDAD DGPPF CIERRE {abr3}"

    # ── Anchos exactos del formato de referencia ──────────────────────
    ws.column_dimensions['A'].width = 9.43
    ws.column_dimensions['B'].width = 7.29
    ws.column_dimensions['C'].width = 67.57
    ws.column_dimensions['D'].width = 15.14
    ws.column_dimensions['E'].width = 15.14
    ws.column_dimensions['F'].width = 15.14
    ws.column_dimensions['G'].width = 20.0
    ws.column_dimensions['H'].width = 11.43

    # ── Altos de fila exactos ─────────────────────────────────────────
    for r, h in {3:19.5, 4:19.5, 5:15.0, 8:15.75, 9:25.5, 10:6.75, 12:6.75}.items():
        ws.row_dimensions[r].height = h

    # ── FILAS 3-4: Títulos (Calibri 15 bold, ha=center, sin fill) ────
    ws.merge_cells('A3:G3')
    ws['A3'].value = 'SECRETARÍA DE AGRICULTURA Y DESARROLLO RURAL'
    ws['A3'].font  = Font(name=FONT, size=15, bold=True)
    ws['A3'].alignment = _a('center', 'center')

    ws.merge_cells('A4:G4')
    ws['A4'].value = 'CONCEPTOS SUJETOS A AUSTERIDAD Y RACIONALIDAD'
    ws['A4'].font  = Font(name=FONT, size=15, bold=True)
    ws['A4'].alignment = _a('center', 'center')

    # ── FILAS 5-6: Subtítulo (Calibri 11 bold, ha=center, wrap) ──────
    ws.merge_cells('A5:G6')
    ws['A5'].value = ('(Artículo 10 Ley Federal de Austeridad Republicana y numeral '
                      '10 fracción II de \nlos Lineamientos en Materia de Austeridad '
                      'Republicana de la Administración Pública Federal)')
    ws['A5'].font  = Font(name=FONT, size=SZ, bold=True)
    ws['A5'].alignment = _a('center', 'center', wrap=True)

    # ── FILAS 8-9: Encabezados de tabla ──────────────────────────────
    # Merges exactos del formato:
    ws.merge_cells('A8:A9')
    ws.merge_cells('B8:B9')
    ws.merge_cells('C8:C9')
    ws.merge_cells('D8:E8')
    ws.merge_cells('F8:F9')
    ws.merge_cells('G8:G9')

    _set_enc(ws['A8'], 'Concepto',             'center', 'center')
    _set_enc(ws['B8'], 'Partida',              'center', 'center')
    _set_enc(ws['C8'], 'Denominación partida', 'center', 'center')
    _set_enc(ws['D8'], 'Ejercido*',            'center', 'center')  # D8:E8 merged
    _set_enc(ws['F8'], 'Diferencia',           'center', 'center', FMT_NUM)
    _set_enc(ws['G8'], f'  {anio} vs {anio-1} (variación porcentual)',
             'center', 'center', FMT_NUM, wrap=True)
    _set_enc(ws['D9'], f'{anio-1}/1',          'center')
    _set_enc(ws['E9'], f'{anio}/2',            'center')

    # ── FILA 11: Total general ────────────────────────────────────────
    # Merge A11:C11 NO existe en el referencia — A11 solo con valor, B y C vacíos
    ws['A11'].value = 'Total general'
    ws['A11'].font  = Font(name=FONT, size=SZ, bold=True)
    ws['B11'].font  = Font(name=FONT, size=SZ, bold=True)
    ws['C11'].font  = Font(name=FONT, size=SZ, bold=True)

    refs_d = '+'.join(f'D{r}' for r in FILAS_GRUPO)
    refs_e = '+'.join(f'E{r}' for r in FILAS_GRUPO)
    ws['D11'].value = f'=+{refs_d}'
    ws['E11'].value = f'=+{refs_e}'
    ws['F11'].value = '=+D11-E11'
    ws['G11'].value = '=(E11-D11)/D11'
    for col in ['D','E','F']:
        c_ = ws[f'{col}11']
        c_.font = Font(name=FONT, size=SZ, bold=True)
        c_.alignment = _a('center')
        c_.number_format = FMT_NUM
    ws['G11'].font = Font(name=FONT, size=SZ, bold=True)
    ws['G11'].alignment = _a('center')
    ws['G11'].number_format = FMT_PCT

    # ── GRUPOS Y PARTIDAS ─────────────────────────────────────────────
    # Llevar registro de qué filas ya se usaron para 32301
    # (fila 20 = Arrendamientos, fila 42 = Fotocopiado)
    # Para fila 42 usaremos el mismo valor de P5=32301 — el usuario ajusta si tiene split
    _32301_ya_usado = False

    for fila, tipo, label_a, partida, fml_d, fml_e in MAPA:
        if tipo in ('grupo', 'grupo1', 'grupof'):
            # Fila de concepto: bold, sin fill, va=top
            ws[f'A{fila}'].value = label_a
            ws[f'A{fila}'].font  = _f(bold=True)
            ws[f'A{fila}'].alignment = _a('left', 'top')

            _set_grupo_num(ws[f'D{fila}'], fml_d)
            _set_grupo_num(ws[f'E{fila}'], fml_e)
            # F y G: fórmula diferencia y variación
            ws[f'F{fila}'].value = f'=+D{fila}-E{fila}'
            ws[f'F{fila}'].font  = _f(bold=True)
            ws[f'F{fila}'].alignment = _a('center', 'top')
            ws[f'F{fila}'].number_format = FMT_NUM
            # G no siempre está en el ref para grupo1 (Congresos no tiene G)
            if tipo != 'grupo1' or label_a not in ('Congresos',):
                ws[f'G{fila}'].value = f'=(E{fila}-D{fila})/D{fila}'
                ws[f'G{fila}'].font  = _f(bold=True)
                ws[f'G{fila}'].alignment = _a('center', 'top')
                ws[f'G{fila}'].number_format = FMT_PCT

        elif tipo == 'part':
            # B: número de partida — ha=center excepto fotocopiado (42,43) que es ha=left
            ws[f'B{fila}'].value = partida
            ws[f'B{fila}'].font  = _f()
            ws[f'B{fila}'].alignment = _a('left' if fila in (42,43) else 'center', 'top')

            # C: denominación
            ws[f'C{fila}'].value = DENOMINACIONES.get(partida, '')
            ws[f'C{fila}'].font  = _f()
            ws[f'C{fila}'].alignment = _a('left', 'top', wrap=True)

            # D y E: valores con regla de split por fila
            # Fila 20 = 32301 Arrendamientos (UR 513)
            # Fila 42 = 32301 Fotocopiado (URs != 513)
            # Fila 43 = 33602 Fotocopiado (UR 512)
            if fila == 20:
                key25 = ('32301', 'arrend')
                key26 = ('32301', 'arrend')
            elif fila == 42:
                key25 = ('32301', 'foto')
                key26 = ('32301', 'foto')
            elif fila == 43:
                key25 = ('33602', 'foto')
                key26 = ('33602', 'foto')
            else:
                key25 = partida
                key26 = partida

            v25 = ej25.get(key25, 0.0)
            v26 = ej26.get(key26, 0.0)

            _set_part_num(ws[f'D{fila}'], v25)
            _set_part_num(ws[f'E{fila}'], v26)

            # F: fórmula diferencia
            ws[f'F{fila}'].value = f'=+D{fila}-E{fila}'
            ws[f'F{fila}'].font  = _f()
            ws[f'F{fila}'].alignment = _a('center', 'top')
            ws[f'F{fila}'].number_format = FMT_NUM

            # G: variación %, con IFERROR para partidas en cero
            ws[f'G{fila}'].value = f'=IFERROR((E{fila}-D{fila})/D{fila},0)'
            ws[f'G{fila}'].font  = _f()
            ws[f'G{fila}'].alignment = _a('center', 'top')
            ws[f'G{fila}'].number_format = FMT_PCT

    # Partidas que el ref tiene G vacío (ceros sin fórmula porcentual)
    # -> ya manejado con IFERROR

    # ── NOTAS AL PIE (posición exacta: filas 69-77) ───────────────────
    ws['A69'].value = 'Nota: '
    ws['A69'].font  = _f(bold=True)
    ws['A69'].alignment = _a('left', 'top')

    # R70: * texto
    ws['A70'].value = '*'
    ws['A70'].font  = _f()
    ws['A70'].alignment = _a('right', 'top')
    ws['B70'].value = ('La información presentada considera Sector Central, '
                       'incluyendo las Oficinas de Representación en las Entidades Federativas.')
    ws['B70'].font  = _f()
    ws['B70'].alignment = _a('left', 'top')

    # R71: texto (sin asterisco en col A)
    ws['B71'].value = ('La información referente al servicio de Fotocopiado deberá ser '
                       'solicitada a la Dirección General de Recursos Materiales, '
                       'Inmuebles y Servicios (Ur 512)')
    ws['B71'].font  = _f()
    ws['B71'].alignment = _a('left', 'top')

    # R72: ** texto largo (merged B72:G72)
    ws['A72'].value = '**'
    ws['A72'].font  = _f()
    ws['A72'].alignment = _a('right', 'top')
    ws.merge_cells('B72:G72')
    ws['B72'].value = ('Para el caso de la partidas 32301 "Arrendamiento de equipo y '
                       'bienes informáticos" y 33602 "Otros servicios comerciales" no es '
                       'posible desagregar de los sistemas hacendarios el gasto del servicio '
                       'de Fotocopiado, toda vez que en estas partidas se cubren diversas '
                       'erogaciones, por lo que la información debiera corresponder a lo '
                       'reportado por las Unidades Responsables ejecutoras del gasto '
                       '(Dirección General de Recursos Materiales, Inmuebles y Servicios '
                       'y las OREF).')
    ws['B72'].font  = _f()
    ws['B72'].alignment = _a('left', 'top', wrap=True)

    # R75: Fuente
    ws['A75'].value = 'Fuente:'
    ws['A75'].font  = _f(bold=True)
    ws['A75'].alignment = _a('left', 'top')

    ws['B76'].value = (f'/1. Sistema de Contabilidad y Presupuesto del ejercicio '
                       f'{anio-1} (base para Cuenta Pública).')
    ws['B76'].font  = _f()
    ws['B76'].alignment = _a('left', 'top')

    ws['B77'].value = (f'/2. Sistema de Contabilidad y Presupuesto del ejercicio '
                       f'{anio}, corte al {d} de {m_es} de {anio}.')
    ws['B77'].font  = _f()
    ws['B77'].alignment = _a('left', 'top')

    # ── Guardar ───────────────────────────────────────────────────────
    if ruta_salida is None:
        ruta_salida = os.path.join(os.getcwd(),
            f"FORMATO_AUSTERIDAD_{abr3}-{str(anio)[2:]}.xlsx")
    wb.save(ruta_salida)
    print(f"\n Archivo generado: {ruta_salida}")
    return ruta_salida

# ══════════════════════════════════════════════════════════════════════
#  COLAB / CLI
# ══════════════════════════════════════════════════════════════════════

def _colab_main():
    try:
        from google.colab import files as _f_
    except ImportError:
        args = _sys.argv[1:]
        if len(args) >= 2:
            generar_austeridad(args[0], args[1])
        else:
            cp = input('Ruta CSV Cuenta Pública 2025: ').strip()
            sc = input('Ruta CSV SICOP corte 2026: ').strip()
            if cp and sc: generar_austeridad(cp, sc)
        return

    print('═'*55)
    print('  FORMATO AUSTERIDAD — SADER')
    print('═'*55)
    print('\n 1/2 — Sube el CSV de CUENTA PÚBLICA cierre 2025:')
    sub1 = _f_.upload()
    if not sub1: return
    n1, c1 = next(iter(sub1.items()))
    r1 = f'/content/{n1}'
    with open(r1,'wb') as f: f.write(c1)
    print(f'    {n1}')
    print('\n 2/2 — Sube el CSV del SICOP corte actual 2026:')
    sub2 = _f_.upload()
    if not sub2: return
    n2, c2 = next(iter(sub2.items()))
    r2 = f'/content/{n2}'
    with open(r2,'wb') as f: f.write(c2)
    print(f'    {n2}')
    rout = generar_austeridad(r1, r2)
    print('\n⬇️  Descargando...')
    _f_.download(rout)
    print(' ¡Listo!')

if __name__ == '__main__':
    _colab_main()
